import urllib
import logging
import configparser
from openpyxl import load_workbook
from Strava import token_exchange
from TrainingSpreadsheet import Calandar, get_training_plan
from flask import Flask, request, session, redirect, render_template
from flask_session import Session


logging.basicConfig(filename="app.log", level=logging.DEBUG)

config = configparser.ConfigParser()
config.read("config.ini")

workbook = load_workbook(config['training plan']['Spreadsheet'])

app = Flask(__name__)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.secret_key = config['flask']['SecretKey']
Session(app)


@app.route("/", methods=['GET'])
def index():
    return render_template("Index.html")


# Redirect to Strava OAuth
@app.route("/authorize", methods=['GET'])
def authorize_strava():
    params = {
        'client_id': config['strava']['ClientID'],
        'redirect_uri': "http://jordancameron.co.uk/token",
        'response_type': "code",
        'scope': "read_all,profile:read_all,activity:read_all"
    }
    return redirect('https://www.strava.com/oauth/authorize?{}'.format(urllib.parse.urlencode(params)))


# Run token exchange to get athletes auth info
@app.route("/token", methods=['GET'])
def token():
    code = request.args.get('code')
    # Todo: Store auth somewhere
    id, auth = token_exchange(config['strava']['ClientID'], config['strava']['ClientSecret'], code)
    session[id] = auth
    return redirect('/')


# Temporary path to expose auth stored in session for testing api calls outside of this app
@app.route("/<id>/auth", methods=['GET'])
def athlete_auth(id):
    auth = session[id]
    return f"{auth}"


# Show activities associated with the athlete with id
@app.route("/<id>/activities", methods=['GET'])
def athlete_activities(id):
    return ""


# Render the list of training plans from the loaded spreadsheet
@app.route("/training", methods=['GET'])
def training_spreadsheet():
    return render_template("Training.html", training_plan_list=workbook.sheetnames)


# Render the training plan from the named sheet
@app.route("/training/<name>", methods=['GET'])
def training_plan(name):
    training_plan = get_training_plan(workbook, name)
    return render_template("Calandar.html", calandar=Calandar(training_plan))
