from openpyxl import load_workbook
from datetime import timedelta

class Run:
    
    def __init__(self, date, run_type=None, plan=None, actual_distance=None):
        self.date = date
        self.run_type = run_type
        self.plan_distance = plan
        self.actual_distance = actual_distance
        
    def __str__(self):
        run = "Type: {} Plan:{} Actual:{}\n".format(self.run_type, self.plan_distance, self.actual_distance)
        return run


# A full calandar representation of the runs, used for html templating
class Calandar:

        def __init__(self, training_plan):
            self.run_weeks = []
            # Iterate over training plan range
            delta = timedelta(days=0)
            run_week = []
            while (date := training_plan.start_date + delta) <= training_plan.end_date:
                run_found = False
                for run in training_plan.runs:
                    if date == run.date:
                        run_found = True
                        run_week.append(run)
                if not run_found:
                    run_week.append(Run(date))
                if (delta.days + 1) % 7 == 0:
                    self.run_weeks.append(run_week)
                    run_week = []
                delta += timedelta(days=1)


class TrainingPlan:

    # Constructor that creates an empty training plan
    def __init__(self):
        self.name = None
        self.start_date = None
        self.end_date = None
        self.runs = []

    # Fill the object with pre-existing training plan data
    def load_plan(self, worksheet):
        self.name = worksheet.title
        self.start_date = worksheet['C4'].value
        # Find the end date on the training plan
        row_num = 4
        while True:
            if worksheet.cell(row=row_num, column=4).value != None:
                row_num += 1
            else:
                row_num -= 1
                break
        self.end_date = worksheet.cell(row=row_num, column=4).value
        # Iterate over each potential date in the training plan
        delta = timedelta(days=0)
        row_num = 4
        col_num = 5
        while (date := self.start_date + delta) <= self.end_date:
            run_type = worksheet.cell(row=row_num, column=col_num).value
            plan = worksheet.cell(row=row_num, column=col_num + 1).value
            actual = worksheet.cell(row=row_num, column=col_num + 2).value
            # Check if end of week period and reset col_num and increase row_num
            if (delta.days + 1) % 7 == 0:
                col_num = 5
                row_num += 1
            else:
                col_num += 3
            if run_type!=None:
                self.runs.append(Run(date, run_type, plan, actual))
            delta += timedelta(days=1)


    def __str__(self):
        training_plan = "" 
        delta = timedelta(days=0)
        while (date := self.start_date + delta) <= self.end_date:
            run_found = False
            for run in self.runs:
                if date == run.date:
                    run_found = True
                    training_plan = training_plan + "{} ".format(date) + str(run)
            if not run_found:
                training_plan = training_plan + "{}\n".format(date)
            delta += timedelta(days=1)
        return training_plan 


# Function to return the specified training plan object from one of the worksheets in workbook
def get_training_plan(workbook, name):
    worksheet = workbook[name]
    training_plan = TrainingPlan()
    training_plan.load_plan(worksheet)
    return training_plan


#================================================
# Test Rig
#================================================
if __name__ == "__main__":
    workbook_name = "RunTraining.xlsx"
    worksheet_name = "Feb 10k"
    workbook = load_workbook(workbook_name)
    # Check if training plan already exists
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
        training_plan = TrainingPlan()
        training_plan.load_plan(worksheet)
        calandar = Calandar(training_plan)
    else:
        training_plan = TrainingPlan()
